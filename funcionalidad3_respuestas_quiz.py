import requests
import json
import re
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.utils import get_column_letter

# === CONFIGURACIÓN ===
token = "934a5bc65d092299e862902196a6f43b"
domain = "https://campusvirtual-sifods.minedu.gob.pe"
restformat = "json"

userid = 423   # <<-- pon aquí el id del usuario
courseid = 1483  # <<-- pon aquí el id del curso
quizid = 1530    # <<-- pon aquí el id del quiz

# === PRIMER PASO: obtener intentos del usuario ===
function = "mod_quiz_get_user_attempts"
params = {
    "wstoken": token,
    "wsfunction": function,
    "moodlewsrestformat": restformat,
    "quizid": quizid,
    "userid": userid,
    "status": "all"   # así traemos todos, no solo finalizados
}

resp = requests.get(f"{domain}/webservice/rest/server.php", params=params)
data_attempts = resp.json()

# === SEGUNDO PASO: recorrer los intentos y obtener sus preguntas ===
rows = []
for intento in data_attempts.get("attempts", []):
    attemptid = intento["id"]
    numero_intento = intento["attempt"]

    # pedir detalle del intento
    function = "mod_quiz_get_attempt_review"
    params = {
        "wstoken": token,
        "wsfunction": function,
        "moodlewsrestformat": restformat,
        "attemptid": attemptid,
        "page": -1
    }
    resp = requests.get(f"{domain}/webservice/rest/server.php", params=params)
    data_review = resp.json()

    for q in data_review.get("questions", []):
        numero = q.get("number", "N/A")

        # Enunciado limpio
        enunciado_html = q.get("html", "")
        soup = BeautifulSoup(enunciado_html, "html.parser")

        # Normalmente Moodle envuelve el enunciado en div.questiontext
        enunciado_div = soup.find("div", class_="qtext")
        if enunciado_div:
            enunciado = enunciado_div.get_text(" ", strip=True)
        else:
            enunciado = soup.get_text(" ", strip=True)[:300]  # fallback


        # Respuesta correcta
        respuesta_correcta = ""
        right_ans = soup.find("div", class_="rightanswer")
        if right_ans:
            respuesta_correcta = right_ans.get_text(" ", strip=True).replace("La respuesta correcta es:", "").strip()

        # Respuesta del estudiante (a veces viene en el html general)
        # Respuesta del estudiante → buscamos "Respuesta guardada:" en el historial
        respuesta_estudiante = "⛔ No disponible"
        texto_plano = soup.get_text(" ", strip=True)

        # 🔎 Buscar "Guardada:" y cortar hasta antes de la palabra "Respuesta"
        match = re.search(r"Guardada:\s*(.+?)\s*Respuesta", texto_plano)
        if match:
            respuesta_estudiante = match.group(1).strip()
        else:
            # Fallback por si no está la palabra "Respuesta"
            if "Guardada:" in texto_plano:
                respuesta_estudiante = texto_plano.split("Guardada:")[-1].strip()

        puntaje = q.get("mark", 0)
        maximo = q.get("maxmark", 0)

        rows.append([
            numero_intento,  # ← número de intento
            numero,
            enunciado,
            respuesta_estudiante,
            respuesta_correcta,
            puntaje,
            maximo
        ])

# === CREACIÓN DEL EXCEL ===
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Resultados Quiz"

# Encabezados
headers = ["Intento N°", "Pregunta N°", "Enunciado", "Respuesta estudiante", "Respuesta correcta", "Puntaje obtenido", "Puntaje máximo"]
ws.append(headers)

# Filas
for row in rows:
    ws.append(row)

# Ajustar el ancho de las columnas automáticamente
for col in ws.columns:
    max_length = 0
    column = col[0].column
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[get_column_letter(column)].width = adjusted_width

# Guardamos el archivo
output_path = r"C:\Users\User\Documents\IA\Minedu_IA\IA_Cuestionarios\Curso1_Formador_Formadores\resultados_quiz_todos_intentos.xlsx"
wb.save(output_path)

print(f"✅ Archivo Excel generado: {output_path}")
